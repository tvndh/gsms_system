import json
from urllib import request
import uuid
import random
import string
import logging
from datetime import timedelta
import base64

from django.db.models import ProtectedError
from django.http import JsonResponse
from collections import defaultdict
from django.db.models import Q
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.db import transaction
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import Avg, Count, Q

# === IMPORT MODELS ĐÃ ĐƯỢC CẬP NHẬT (Xóa DanhMucDichVu, SanPhamDichVu) ===
from .models import (
    TramXang, BonChua, NhaCungCap, HoaDon, ChiTietHoaDon, 
    TinTuc, DanhMuc, SanPham, PhieuNhap, YeuCauNhapHang, BangGiaNhienLieu, 
    DanhGiaTram, DoiTacB2B, HoSoUngVien, LienHeGopY
)

User = get_user_model()
logger = logging.getLogger(__name__)

# ==========================================
# 1. HỆ THỐNG XÁC THỰC 
# ==========================================
def dang_nhap(request):
    if request.user.is_authenticated:
        return phan_luong_dashboard(request.user)

    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        
        if user is not None:
            if not user.is_active:
                messages.error(request, "Tài khoản của bạn đã bị khóa!")
                return redirect('login')
                
            login(request, user)
            messages.success(request, f"Xin chào {user.full_name or user.username}!")
            return phan_luong_dashboard(user)
        else:
            messages.error(request, "Tên đăng nhập hoặc mật khẩu không đúng!")
            
    return render(request, 'pages/login.html')

def phan_luong_dashboard(user):
    if user.role in ['admin', 'ke_toan'] or user.is_superuser:
        return redirect('admin_dashboard')
    elif user.role == 'truong_tram':
        return redirect('bao_cao_tram')
    elif user.role == 'nv_ban_xang':
        return redirect('pos_xang')
    else:
        return redirect('trang_chu')

def dang_xuat(request):
    logout(request)
    messages.info(request, "Đăng xuất thành công.")
    return redirect('login')

# ==========================================
# 2. GIAO DIỆN KHÁCH & TRANG CHỦ (PAGES)
# ==========================================
def guest_home(request):
    tin_moi_nhat = TinTuc.objects.order_by('-ngay_dang')[:3]
    san_pham_hot = SanPham.objects.all()[:4]
    
    # BƯỚC 1: Dùng annotate để gộp luôn điểm trung bình và số lượt đánh giá vào Trạm
    trams = TramXang.objects.filter(trang_thai='hoat_dong').annotate(
        diem_danh_gia=Avg('cac_danh_gia__so_sao', filter=Q(cac_danh_gia__da_duyet=True)),
        luot_danh_gia=Count('cac_danh_gia', filter=Q(cac_danh_gia__da_duyet=True))
    )
    
    # BƯỚC 2: ĐÃ SỬA LỖI Ở ĐÂY - Bỏ điều kiện lọc "> 0" 
    # Giờ hệ thống sẽ ghi nhận TẤT CẢ các bồn chứa, kể cả bồn đang trống (0 lít)
    bon_chua_co_xang = BonChua.objects.values_list('tram_id', 'loai_nhien_lieu')

    nhien_lieu_theo_tram = defaultdict(list)
    for tram_id, loai_nl in bon_chua_co_xang:
        nhien_lieu_theo_tram[tram_id].append(loai_nl)

    tram_list = []
    for t in trams:
        # Lấy danh sách nhiên liệu TỪ BỒN CHỨA (Trạm nào có bồn nào thì hiện loại xăng đó)
        available_fuels = nhien_lieu_theo_tram.get(t.id, [])

        # Lấy danh sách đánh giá
        cac_danh_gia = t.cac_danh_gia.filter(da_duyet=True).order_by('-ngay_gui')[:10]
        list_dg = []
        for dg in cac_danh_gia:
            list_dg.append({
                'ten': dg.ten_khach_hang,
                'sao': dg.so_sao,
                'noi_dung': dg.noi_dung,
                'ngay': dg.ngay_gui.strftime("%d/%m/%Y")
            })

        tram_list.append({
            'id': t.id, 
            'ten': t.ten_tram, 
            'lat': float(t.latitude), 
            'lng': float(t.longitude),
            'dia_chi': t.dia_chi, 
            'fuels': available_fuels, # Truyền đúng dữ liệu thật vào bản đồ
            'diem_danh_gia': float(t.diem_danh_gia) if t.diem_danh_gia else 5.0,
            'luot_danh_gia': t.luot_danh_gia,
            'danh_gia_chi_tiet': list_dg 
        })

    # Cập nhật Bảng giá xăng
    gia_db = {gia.loai_nhien_lieu: gia.gia_ban for gia in BangGiaNhienLieu.objects.all()}
    gia_hien_tai = {
        'A95': gia_db.get('A95', 24500), 'E5': gia_db.get('E5', 23500),
        'E10': gia_db.get('E10', 24000), 'DO': gia_db.get('DO', 21000),
    }

    context = {
        'tin_moi_nhat': tin_moi_nhat,
        'san_pham': san_pham_hot,
        'tram_json': json.dumps(tram_list), 
        'gia': gia_hien_tai
    }
    return render(request, 'pages/trang_chu.html', context)
def trang_gioi_thieu(request): return render(request, 'pages/gioi_thieu.html')

def trang_tin_tuc(request):
    tu_khoa = request.GET.get('q', '')
    if tu_khoa:
        ds_tin_goc = TinTuc.objects.filter(
            Q(tieu_de__icontains=tu_khoa) | Q(tom_tat__icontains=tu_khoa)
        ).order_by('-ngay_dang')
    else:
        ds_tin_goc = TinTuc.objects.all().order_by('-ngay_dang')

    paginator = Paginator(ds_tin_goc, 6) 
    page_number = request.GET.get('page')
    ds_tin = paginator.get_page(page_number)

    return render(request, 'pages/tin_tuc.html', {'ds_tin': ds_tin, 'tu_khoa': tu_khoa})

def chi_tiet_tin_tuc(request, id):
    tin = get_object_or_404(TinTuc, id=id)
    tin_lien_quan = TinTuc.objects.exclude(id=id).order_by('-ngay_dang')[:3]
    return render(request, 'pages/chi_tiet_tin_tuc.html', {'tin': tin, 'tin_lien_quan': tin_lien_quan})

def trang_san_pham(request): 
    # Đã gộp toàn bộ sản phẩm/dịch vụ vào 1 bảng duy nhất
    ds_sp = SanPham.objects.select_related('danh_muc').all()
    ds_danh_muc = DanhMuc.objects.all()
    
    return render(request, 'pages/san_pham.html', {
        'ds_sp': ds_sp,
        'ds_danh_muc': ds_danh_muc
    })

def chi_tiet_linh_vuc(request, slug):
    data = {
        'kinh-doanh-xang-dau': {'title': 'Kinh Doanh Xăng Dầu', 'img': '/static/images/banners/kinh-doanh-xang-dau.jpg', 'content': '...', 'cam_ket': []},
        'van-tai-xang-dau': {'title': 'Vận Tải Xăng Dầu', 'img': '/static/images/banners/van-tai-xang-dau.jpg', 'content': '...', 'cam_ket': []},
        'khi-hoa-long': {'title': 'Khí Hóa Lỏng (LPG)', 'img': '/static/images/banners/khi-hoa-long.jpg', 'content': '...', 'cam_ket': []},
        'hoa-dau-dung-moi': {'title': 'Hóa Dầu & Dung Môi', 'img': '/static/images/banners/hoa-dau-dung-moi.jpg', 'content': '...', 'cam_ket': []},
        'dich-vu-tai-chinh': {'title': 'Dịch Vụ Tài Chính', 'img': '/static/images/banners/dich-vu-tai-chinh.jpg', 'content': '...', 'cam_ket': []}
    }
    context = data.get(slug)
    if not context: raise Http404("Không tìm thấy lĩnh vực này")
    return render(request, 'pages/linh_vuc.html', context)

def trang_lien_he(request):
    if request.method == 'POST':
        ho_ten = request.POST.get('ho_ten')
        so_dien_thoai = request.POST.get('so_dien_thoai')
        email_khach = request.POST.get('email')
        tieu_de = request.POST.get('tieu_de')
        noi_dung = request.POST.get('noi_dung')

        try:
            # 1. Lưu vào CSDL
            obj = LienHeGopY.objects.create(
                ho_ten=ho_ten, so_dien_thoai=so_dien_thoai,
                email=email_khach, tieu_de=tieu_de, noi_dung=noi_dung
            )

            # 2. TỰ ĐỘNG PHẢN HỒI "CẢM ƠN" NGAY LẬP TỨC
            if email_khach:
                subject = "GSMS - Cảm ơn bạn đã gửi liên hệ"
                message = f"Chào {ho_ten},\n\nHệ thống GSMS đã nhận được yêu cầu của bạn về '{tieu_de}'. Chúng tôi sẽ phản hồi sớm nhất có thể.\n\nTrân trọng."
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email_khach], fail_silently=True)

            messages.success(request, "Đã gửi thành công! Vui lòng kiểm tra Email.")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
        return redirect('lien_he')
    return render(request, 'pages/lien_he.html')

def trang_doi_tac(request): return render(request, 'pages/doi_tac.html')

def trang_tuyen_dung(request): return render(request, 'pages/tuyen_dung.html')

# ==========================================
# 3. NHÂN VIÊN & TRƯỞNG TRẠM
# ==========================================
@login_required
def pos_xang(request):
    user = request.user
    if user.role in ['admin', 'ke_toan'] or user.is_superuser: 
        return redirect('admin_dashboard')
        
    if not user.tram_xang:
        messages.error(request, "Tài khoản của bạn chưa được phân công về Trạm Xăng nào!")
        return redirect('login')

    tram_cua_toi = user.tram_xang
    ds_bon_raw = BonChua.objects.filter(tram=tram_cua_toi)

    ds_bon = []
    bon_can_canh_bao = []
    for b in ds_bon_raw:
        if b.suc_chua_toi_da > 0 and (b.muc_hien_tai / b.suc_chua_toi_da) * 100 < 20:
            bon_can_canh_bao.append(b)
            
        # Đã xóa dòng gán b.gia_ban_hien_tai ở đây để tránh lỗi AttributeError
        ds_bon.append(b)

    if user.role == 'truong_tram':
        lich_su = HoaDon.objects.filter(nhan_vien__tram_xang=tram_cua_toi).order_by('-thoi_gian')[:20]
    else:
        lich_su = HoaDon.objects.filter(nhan_vien=user).order_by('-thoi_gian')[:10]

    return render(request, 'staff/pos_xang.html', {
        'tram': tram_cua_toi, 
        'ds_bon': ds_bon, 
        'bon_can_canh_bao': bon_can_canh_bao, 
        'lich_su_ban': lich_su
    })

@login_required
def xu_ly_ban_hang(request):
    if request.method == 'POST':
        if request.user.role in ['admin', 'ke_toan'] or request.user.is_superuser: 
            return redirect('admin_dashboard')
            
        try:
            loai_nl = request.POST.get('loai_nhien_lieu')
            so_tien = float(request.POST.get('so_tien'))
            pt_thanh_toan = request.POST.get('phuong_thuc_thanh_toan', 'tien_mat')
            
            gia_db = BangGiaNhienLieu.objects.filter(loai_nhien_lieu=loai_nl).first()
            don_gia = gia_db.gia_ban if gia_db else 20000
            so_lit = so_tien / don_gia

            with transaction.atomic(): 
                bon = BonChua.objects.select_for_update().get(tram=request.user.tram_xang, loai_nhien_lieu=loai_nl)
                if bon.muc_hien_tai >= so_lit:
                    bon.muc_hien_tai -= so_lit
                    bon.save()
                    
                    hd = HoaDon.objects.create(
                        ma_hd=f"HD-{timezone.now().strftime('%y%m%d%H%M%S')}-{request.user.id}", 
                        nhan_vien=request.user, 
                        tong_tien=so_tien,
                        phuong_thuc_thanh_toan=pt_thanh_toan
                    )
                    ChiTietHoaDon.objects.create(hoa_don=hd, ten_mat_hang=f"Nhiên liệu {loai_nl}", so_luong=so_lit, don_gia=don_gia, thanh_tien=so_tien)
                    messages.success(request, f"Thanh toán thành công: {so_lit:.2f} Lít {loai_nl}!")
                else:
                    messages.error(request, "Bồn không đủ nhiên liệu để xuất!")
        except Exception as e:
            messages.error(request, "Có lỗi xảy ra, vui lòng thử lại!")
    return redirect('pos_xang')

@login_required
def staff_chot_ca(request):
    today = timezone.now().date()
    ds_hoa_don = HoaDon.objects.filter(nhan_vien=request.user, thoi_gian__date=today)
    tong_tien = ds_hoa_don.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0
    tong_lit = ChiTietHoaDon.objects.filter(hoa_don__in=ds_hoa_don).aggregate(Sum('so_luong'))['so_luong__sum'] or 0
    
    return render(request, 'staff/staff_chot_ca.html', {'tong_tien': tong_tien, 'so_gd': ds_hoa_don.count(), 'tong_lit': tong_lit, 'ngay_chot': timezone.now()})
@login_required
def tao_yeu_cau_nhap_hang(request):
    if request.method == 'POST':
        if request.user.role != 'truong_tram': return redirect('pos_xang')
        YeuCauNhapHang.objects.create(
            tram=request.user.tram_xang,
            loai_nhien_lieu=request.POST.get('loai_nhien_lieu'),
            so_luong=float(request.POST.get('so_luong')),
            ghi_chu=request.POST.get('ghi_chu', ''),
            trang_thai='cho_duyet'
        )
        messages.success(request, "Đã gửi yêu cầu cấp hàng lên Tổng Công Ty!")
    return redirect('pos_xang')

@login_required
def bao_cao_tram(request):
    if request.user.role != 'truong_tram': return redirect('pos_xang')
    tram = request.user.tram_xang
    today = timezone.now().date()
    hds_hom_nay = HoaDon.objects.filter(nhan_vien__tram_xang=tram, thoi_gian__date=today)
    doanh_thu_nhan_vien = HoaDon.objects.filter(nhan_vien__tram_xang=tram, thoi_gian__date=today).values('nhan_vien__username', 'nhan_vien__full_name').annotate(tong_ban=Sum('tong_tien'), so_don=Count('id')).order_by('-tong_ban')
    
    return render(request, 'staff/bao_cao_tram.html', {
        'tram': tram, 'ngay_bao_cao': timezone.now(),
        'doanh_thu_hom_nay': hds_hom_nay.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0,
        'so_gd_hom_nay': hds_hom_nay.count(),
        'san_luong_hom_nay': ChiTietHoaDon.objects.filter(hoa_don__in=hds_hom_nay).aggregate(Sum('so_luong'))['so_luong__sum'] or 0,
        'ds_bon': BonChua.objects.filter(tram=tram),
        'doanh_thu_nhan_vien': doanh_thu_nhan_vien,
    })

# ==========================================
# 4. QUẢN TRỊ TỔNG QUAN (ADMIN & KẾ TOÁN)
# ==========================================
@login_required
def admin_dashboard(request):
    if request.user.role not in ['admin', 'ke_toan'] and not request.user.is_superuser:
        messages.warning(request, "Bạn không có quyền truy cập!")
        return redirect('trang_chu')

    today = timezone.now().date()
    ds_bon = BonChua.objects.all()

    stats = HoaDon.objects.filter(thoi_gian__date=today).aggregate(
        total_money=Sum('tong_tien'), total_tx=Count('id')
    )
    doanh_thu = stats['total_money'] or 0
    so_giao_dich = stats['total_tx'] or 0
    san_luong = ChiTietHoaDon.objects.filter(hoa_don__thoi_gian__date=today).aggregate(Sum('so_luong'))['so_luong__sum'] or 0

    loi_nhuan_hom_nay = doanh_thu * 0.065

    tu_khoa = request.GET.get('q', '')
    ds_tram = TramXang.objects.all()
    
    if tu_khoa:
        ds_tram = ds_tram.filter(ten_tram__icontains=tu_khoa)

    bang_doanh_thu = []
    for t in ds_tram:
        hds = HoaDon.objects.filter(nhan_vien__tram_xang=t, thoi_gian__date=today)
        dt = hds.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0
        sl = ChiTietHoaDon.objects.filter(hoa_don__in=hds).aggregate(Sum('so_luong'))['so_luong__sum'] or 0
        ln_tram = dt * 0.065
        bang_doanh_thu.append({
            'tram': t, 'doanh_thu': dt, 'loi_nhuan': ln_tram, 'san_luong': sl, 'so_don': hds.count()
        })
    bang_doanh_thu.sort(key=lambda x: x['doanh_thu'], reverse=True)

    now = timezone.now()
    
    def lay_tong_nhien_lieu(hds_queryset):
        chi_tiet = ChiTietHoaDon.objects.filter(hoa_don__in=hds_queryset).values('ten_mat_hang').annotate(tong=Sum('so_luong'))
        fuels = [0, 0, 0, 0] 
        for item in chi_tiet:
            ten = item['ten_mat_hang']
            so_luong = float(item['tong'] or 0)
            if 'A95' in ten: fuels[0] += so_luong
            elif 'E5' in ten: fuels[1] += so_luong
            elif 'E10' in ten: fuels[2] += so_luong
            elif 'DO' in ten: fuels[3] += so_luong
        return fuels

    hds_7days = HoaDon.objects.filter(thoi_gian__date__gte=now.date() - timedelta(days=6))
    day_data = {'labels': [], 'revenue': [], 'volume': [], 'fuels': lay_tong_nhien_lieu(hds_7days)}
    for i in range(6, -1, -1):
        dt = now - timedelta(days=i)
        day_data['labels'].append(dt.strftime("%d/%m"))
        hds = HoaDon.objects.filter(thoi_gian__date=dt.date())
        day_data['revenue'].append(float(hds.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0) / 1000000)
        day_data['volume'].append(float(ChiTietHoaDon.objects.filter(hoa_don__in=hds).aggregate(Sum('so_luong'))['so_luong__sum'] or 0))

    hds_28days = HoaDon.objects.filter(thoi_gian__date__gte=now.date() - timedelta(days=27))
    month_data = {'labels': ['Tuần 1', 'Tuần 2', 'Tuần 3', 'Tuần 4'], 'revenue': [0,0,0,0], 'volume': [0,0,0,0], 'fuels': lay_tong_nhien_lieu(hds_28days)}
    for i in range(28):
        dt = now - timedelta(days=i)
        week_idx = 3 - (i // 7)
        hds = HoaDon.objects.filter(thoi_gian__date=dt.date())
        month_data['revenue'][week_idx] += float(hds.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0) / 1000000
        month_data['volume'][week_idx] += float(ChiTietHoaDon.objects.filter(hoa_don__in=hds).aggregate(Sum('so_luong'))['so_luong__sum'] or 0)

    hds_year = HoaDon.objects.filter(thoi_gian__year=now.year)
    year_data = {'labels': [f'T{i}' for i in range(1, 13)], 'revenue': [0]*12, 'volume': [0]*12, 'fuels': lay_tong_nhien_lieu(hds_year)}
    for hd in hds_year:
        m_idx = hd.thoi_gian.month - 1
        year_data['revenue'][m_idx] += float(hd.tong_tien or 0) / 1000000
        sl_sum = ChiTietHoaDon.objects.filter(hoa_don=hd).aggregate(Sum('so_luong'))['so_luong__sum'] or 0
        year_data['volume'][m_idx] += float(sl_sum)

    quarter_data = {
        'labels': ['Quý 1', 'Quý 2', 'Quý 3', 'Quý 4'],
        'revenue': [sum(year_data['revenue'][0:3]), sum(year_data['revenue'][3:6]), sum(year_data['revenue'][6:9]), sum(year_data['revenue'][9:12])],
        'volume': [sum(year_data['volume'][0:3]), sum(year_data['volume'][3:6]), sum(year_data['volume'][6:9]), sum(year_data['volume'][9:12])],
        'fuels': year_data['fuels'] 
    }

    chart_data = {'day': day_data, 'month': month_data, 'quarter': quarter_data, 'year': year_data}
    
    context = {
        'ds_bon': ds_bon,
        'doanh_thu_hom_nay': doanh_thu,
        'loi_nhuan_hom_nay': loi_nhuan_hom_nay, 
        'san_luong_hom_nay': san_luong,
        'so_giao_dich': so_giao_dich,
        'chart_data_json': json.dumps(chart_data),
        'bang_doanh_thu': bang_doanh_thu,
        'tu_khoa': tu_khoa,
    }
    return render(request, 'admin/admin_dashboard.html', context)

@login_required
def admin_import(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Bạn không có quyền truy cập trang nhập hàng!")
        return redirect('trang_chu')

    if request.method == 'POST':
        try:
            ncc_id = request.POST.get('ncc_id')
            danh_sach_bon = request.POST.getlist('bon_nhan[]')
            danh_sach_so_lit = request.POST.getlist('so_luong[]')
            so_km_tong = float(request.POST.get('khoang_cach_tong', 0))
            
            if not ncc_id or not danh_sach_bon or not danh_sach_so_lit:
                messages.error(request, "Vui lòng chọn đầy đủ Kho cung cấp và ít nhất 1 điểm giao!")
                return redirect('admin_import')
                
            so_diem_giao = len(danh_sach_bon)
            tien_cuoc_tong = so_km_tong * 15000
            tien_cuoc_moi_diem = tien_cuoc_tong / so_diem_giao if so_diem_giao > 0 else 0
            tong_lit_da_xuat = 0

            with transaction.atomic():
                ncc = NhaCungCap.objects.select_for_update().get(id=ncc_id) 
                
                for i in range(so_diem_giao):
                    bon_id = danh_sach_bon[i]
                    so_lit = float(danh_sach_so_lit[i])
                    
                    bon = BonChua.objects.select_for_update().get(id=bon_id)
                    loai_nl = bon.loai_nhien_lieu
                    ton_kho_hien_tai = getattr(ncc, f'ton_kho_{loai_nl}', 0) 
                    
                    if ton_kho_hien_tai < so_lit:
                        raise ValueError(f"LỖI: Kho {ncc.ten_ncc} chỉ còn {ton_kho_hien_tai:,.0f} Lít {loai_nl}. Không đủ xuất cho bồn {bon.ten_bon}!")

                    if bon.muc_hien_tai + so_lit > bon.suc_chua_toi_da:
                        raise ValueError(f"CẢNH BÁO: Bồn {bon.ten_bon} tại {bon.tram.ten_tram} sẽ bị tràn. Vui lòng giảm số lượng!")

                    setattr(ncc, f'ton_kho_{loai_nl}', ton_kho_hien_tai - so_lit)
                    ncc.save() 

                    bon.muc_hien_tai += so_lit
                    bon.save() 
                    
                    tong_lit_da_xuat += so_lit
                    gia_nhap_si = 21000 
                    tien_hang = so_lit * gia_nhap_si

                    PhieuNhap.objects.create(
                        ma_pn=f"PN-{timezone.now().strftime('%d%m%H%M%S')}-{random.randint(10,99)}-{i}",
                        nha_cung_cap_id=ncc_id,
                        bon_chua=bon,
                        so_lit_nhap=so_lit,
                        gia_nhap_1_lit=gia_nhap_si,
                        cuoc_van_chuyen=tien_cuoc_moi_diem,
                        tong_chi_phi=tien_hang + tien_cuoc_moi_diem,
                        thanh_tien=tien_hang
                    )

                    YeuCauNhapHang.objects.filter(
                        tram=bon.tram,
                        loai_nhien_lieu=loai_nl,
                        trang_thai='cho_duyet'
                    ).update(trang_thai='da_duyet')

            messages.success(request, f"Đã phát lệnh điều phối thành công! Tổng xuất: {tong_lit_da_xuat:,.0f} lít cho {so_diem_giao} trạm.")
            return redirect('admin_dashboard')
            
        except ValueError as ve:
            messages.error(request, str(ve))
            return redirect('admin_import')
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
            return redirect('admin_import')

    ds_ncc = NhaCungCap.objects.all()
    ncc_list = [{'id': n.id, 'name': n.ten_ncc, 'lat': float(n.latitude or 0), 'lng': float(n.longitude or 0), 'address': n.dia_chi} for n in ds_ncc]

    ds_tram_all = TramXang.objects.all()
    station_list = [{'id': t.id, 'name': t.ten_tram, 'lat': float(t.latitude or 0), 'lng': float(t.longitude or 0)} for t in ds_tram_all]

    ds_bon = BonChua.objects.select_related('tram').all()
    tank_list = []
    
    for b in ds_bon:
        phan_tram = round((b.muc_hien_tai / b.suc_chua_toi_da) * 100, 1) if b.suc_chua_toi_da > 0 else 0
        tank_list.append({
            'id': b.id, 
            'ten_bon': b.ten_bon, 
            'loai': b.get_loai_nhien_lieu_display(),
            'muc_hien_tai': float(b.muc_hien_tai), 
            'suc_chua': float(b.suc_chua_toi_da),
            'phan_tram': phan_tram, 
            'tram_id': b.tram.id
        })

    ds_yeu_cau = YeuCauNhapHang.objects.filter(trang_thai='cho_duyet').order_by('-id')

    context = {
        'ds_ncc': ds_ncc, 
        'ncc_json': json.dumps(ncc_list), 
        'tank_json': json.dumps(tank_list),
        'station_json': json.dumps(station_list), 
        'ds_yeu_cau': ds_yeu_cau,
    }
    
    return render(request, 'admin/admin_import.html', context)

@login_required
def duyet_yeu_cau(request, yc_id):
    if request.user.role != 'admin' and not request.user.is_superuser: return redirect('trang_chu')
    try:
        yc = YeuCauNhapHang.objects.get(id=yc_id)
        yc.trang_thai = 'da_giao'
        yc.save()
        messages.success(request, f"Đã xử lý xong yêu cầu của {yc.tram.ten_tram}!")
    except: 
        messages.error(request, "Không tìm thấy yêu cầu này!")
    return redirect('admin_import')

@login_required
def quan_ly_gia(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    if request.method == 'POST':
        try:
            for loai, key in [('A95', 'gia_A95'), ('E5', 'gia_E5'), ('E10', 'gia_E10'), ('DO', 'gia_DO')]:
                if request.POST.get(key): 
                    BangGiaNhienLieu.objects.update_or_create(loai_nhien_lieu=loai, defaults={'gia_ban': float(request.POST.get(key))})
            messages.success(request, "Đã đồng bộ giá cho toàn hệ thống!")
        except Exception as e: 
            messages.error(request, f"Lỗi: {e}")
        return redirect('quan_ly_gia')
        
    gia_hien_tai = {loai: BangGiaNhienLieu.objects.filter(loai_nhien_lieu=loai).first() for loai in ['A95', 'E5', 'E10', 'DO']}
    return render(request, 'admin/quan_ly_gia.html', {'gia_hien_tai': gia_hien_tai})

@login_required
def xuat_excel_doanh_thu(request):
    if request.user.role not in ['admin', 'ke_toan'] and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        messages.error(request, "Vui lòng cài openpyxl (pip install openpyxl)")
        return redirect('admin_dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bao_Cao_Doanh_Thu_Tram"
    ws.append(['STT', 'Tên Trạm Xăng', 'Địa Chỉ', 'Doanh Thu (VNĐ)', 'Sản Lượng (Lít)', 'Số Đơn'])
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    tu_khoa = request.GET.get('q', '')
    ds_tram = TramXang.objects.filter(ten_tram__icontains=tu_khoa) if tu_khoa else TramXang.objects.all()
    today = timezone.now().date()
    tong_dt = 0

    for stt, t in enumerate(ds_tram, 1):
        hds = HoaDon.objects.filter(nhan_vien__tram_xang=t, thoi_gian__date=today)
        dt = hds.aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0
        sl = ChiTietHoaDon.objects.filter(hoa_don__in=hds).aggregate(Sum('so_luong'))['so_luong__sum'] or 0
        tong_dt += dt
        ws.append([stt, t.ten_tram, t.dia_chi, dt, sl, hds.count()])

    ws.append(['', 'TỔNG CỘNG TOÀN HỆ THỐNG', '', tong_dt, '', ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FF0000") 
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="BaoCao_DoanhThu_GSMS_{timezone.now().strftime("%d%m%Y")}.xlsx"'
    wb.save(response)
    return response

# ==========================================
# 5. CRUD: TRẠM XĂNG
# ==========================================
# ==========================================
# 11. CRUD: TRẠM XĂNG (Đã fix lỗi Bồn chứa)
# ==========================================
import random, string
from django.contrib.auth import get_user_model
User = get_user_model()

@login_required
def admin_danh_sach_tram(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
    ds_tram = TramXang.objects.all().order_by('-id')
    return render(request, 'admin/admin_trams.html', {'ds_tram': ds_tram})

@login_required
def admin_them_tram(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
    
    if request.method == 'POST':
        try:
            # Tự động thay phẩy thành chấm cho tọa độ
            lat_str = str(request.POST.get('latitude', '0')).replace(',', '.')
            lng_str = str(request.POST.get('longitude', '0')).replace(',', '.')

            tram = TramXang.objects.create(
                ten_tram=request.POST.get('ten_tram'), 
                dia_chi=request.POST.get('dia_chi'),
                latitude=float(lat_str), 
                longitude=float(lng_str),
                trang_thai=request.POST.get('trang_thai', 'hoat_dong')
            )
            
            # --- XỬ LÝ TẠO BỒN TỰ ĐỘNG ---
            nhien_lieu_duoc_chon = request.POST.getlist('nhien_lieu')
            thong_so = {'A95': 15000, 'E5': 10000, 'E10': 10000, 'DO': 20000}
            
            for nl in nhien_lieu_duoc_chon:
                if nl in thong_so:
                    # Thay vì create, mình dùng get_or_create cho chắc ăn
                    BonChua.objects.get_or_create(
                        tram=tram, 
                        loai_nhien_lieu=nl,
                        defaults={
                            'ten_bon': f"Bồn {nl}", 
                            'suc_chua_toi_da': thong_so[nl], 
                            'muc_hien_tai': 0
                        }
                    )

            # --- TẠO TÀI KHOẢN TỰ ĐỘNG CÓ RANDOM MẬT KHẨU ---
            p_truong = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
            p_nv = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
            
            User.objects.create_user(username=f"truongtram_{tram.id}", password=p_truong, full_name=f"Trưởng {tram.ten_tram}", role="truong_tram", tram_xang_id=tram.id)
            User.objects.create_user(username=f"nhanvien_{tram.id}", password=p_nv, full_name=f"NV {tram.ten_tram}", role="nv_ban_xang", tram_xang_id=tram.id)

            messages.success(request, f"Đã tạo Trạm! Vui lòng lưu lại MK Trưởng: {p_truong} | NV: {p_nv}")
            return redirect('admin_trams')
            
        except Exception as e:
            messages.error(request, f"Lỗi tạo Trạm: {e}")

    return render(request, 'admin/admin_tram_form.html')

@login_required
def admin_sua_tram(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        messages.error(request, "Bạn không có quyền truy cập!")
        return redirect('trang_chu')
        
    tram = get_object_or_404(TramXang, id=id)
    ds_bon = BonChua.objects.filter(tram=tram)
    
    # Lấy danh sách loại nhiên liệu mà trạm ĐANG CÓ để truyền ra giao diện (đánh dấu checked)
    loai_nhien_lieu_hien_tai = ds_bon.values_list('loai_nhien_lieu', flat=True)
    
    if request.method == 'POST':
        try:
            tram.ten_tram = request.POST.get('ten_tram')
            tram.dia_chi = request.POST.get('dia_chi')
            tram.trang_thai = request.POST.get('trang_thai', 'hoat_dong')
            
            lat_str = str(request.POST.get('latitude', '0')).replace(',', '.')
            lng_str = str(request.POST.get('longitude', '0')).replace(',', '.')
            
            tram.latitude = float(lat_str)
            tram.longitude = float(lng_str)
            tram.save()
            
            # --- ĐÂY LÀ PHẦN SỬA LỖI TỰ ĐẺ BỒN ---
            nhien_lieu_cap_nhat = request.POST.getlist('nhien_lieu')
            thong_so = {'A95': 15000, 'E5': 10000, 'E10': 10000, 'DO': 20000}
            
            # 1. Quét danh sách được tích: Nếu chưa có thì đẻ bồn mới (Dùng get_or_create)
            for nl in nhien_lieu_cap_nhat:
                if nl in thong_so:
                    BonChua.objects.get_or_create(
                        tram=tram, 
                        loai_nhien_lieu=nl,
                        defaults={
                            'ten_bon': f"Bồn {nl}", 
                            'suc_chua_toi_da': thong_so[nl], 
                            'muc_hien_tai': 0
                        }
                    )
            # LƯU Ý BẢO MẬT: Mình KHÔNG code lệnh xóa các bồn bị "bỏ tích", 
            # vì lỡ trong bồn cũ đang chứa 5000 Lít xăng mà bị xóa cái rụp là mất trắng tài sản. 
            # (Admin muốn xóa bồn thì phải tự qua Tab "Quản lý Bồn Chứa" để xóa thủ công).

            messages.success(request, f'Đã cập nhật thông tin {tram.ten_tram} thành công!')
            return redirect('admin_trams')
            
        except Exception as e:
            messages.error(request, f"Lỗi cập nhật: {e}")
        
    context = {
        'tram': tram, 
        'ds_bon': ds_bon,
        'loai_nhien_lieu_hien_tai': loai_nhien_lieu_hien_tai # Truyền cái này ra để giao diện biết ô nào cần Tick
    }
    return render(request, 'admin/admin_tram_form.html', context)

@login_required
def admin_xoa_tram(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    tram = get_object_or_404(TramXang, id=id)
    
    # Bọc try...except phòng trường hợp trạm có dính đến hóa đơn xuất
    try:
        ten_tram = tram.ten_tram
        tram.delete()
        messages.success(request, f'Đã xóa vĩnh viễn {ten_tram} và toàn bộ bồn chứa/tài khoản nhân viên trực thuộc!')
    except Exception as e:
        messages.error(request, "Không thể xóa Trạm xăng này vì đang vướng dữ liệu lịch sử Hóa đơn/Nhập hàng! Hãy đổi trạng thái sang 'Đóng Cửa'.")
        
    return redirect('admin_trams')

# ==========================================
# 6. CRUD: KHO TỔNG / NHÀ CUNG CẤP
# ==========================================
@login_required
def admin_danh_sach_kho(request):
    if request.user.role != 'admin' and not request.user.is_superuser: return redirect('trang_chu')
    ds_kho = NhaCungCap.objects.all().order_by('-id')
    return render(request, 'admin/admin_khos.html', {'ds_kho': ds_kho})

@login_required
def admin_them_kho(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    if request.method == 'POST':
        # Bắt dữ liệu và tự động đổi dấu phẩy (,) thành dấu chấm (.)
        lat_raw = request.POST.get('latitude', '0').replace(',', '.')
        lng_raw = request.POST.get('longitude', '0').replace(',', '.')
        
        try:
            NhaCungCap.objects.create(
                ten_ncc=request.POST.get('ten_ncc'), 
                sdt=request.POST.get('sdt'), 
                dia_chi=request.POST.get('dia_chi'),
                # Ép kiểu an toàn sang float
                latitude=float(lat_raw) if lat_raw else 0.0, 
                longitude=float(lng_raw) if lng_raw else 0.0
            )
            messages.success(request, "Thêm Kho tổng thành công!")
            return redirect('admin_khos')
        except ValueError:
            messages.error(request, "Lỗi: Tọa độ nhập vào không hợp lệ. Vui lòng kiểm tra lại số liệu!")
            
    return render(request, 'admin/admin_kho_form.html')


@login_required
def admin_sua_kho(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    kho = get_object_or_404(NhaCungCap, id=id)
    if request.method == 'POST':
        kho.ten_ncc = request.POST.get('ten_ncc')
        kho.sdt = request.POST.get('sdt')
        kho.dia_chi = request.POST.get('dia_chi')
        
        # Bắt dữ liệu và tự động đổi dấu phẩy (,) thành dấu chấm (.)
        lat_raw = request.POST.get('latitude', '')
        lng_raw = request.POST.get('longitude', '')
        
        try:
            if lat_raw: 
                kho.latitude = float(lat_raw.replace(',', '.'))
            if lng_raw: 
                kho.longitude = float(lng_raw.replace(',', '.'))
                
            kho.save()
            messages.success(request, "Cập nhật Kho thành công!")
            return redirect('admin_khos')
        except ValueError:
            messages.error(request, "Lỗi: Tọa độ nhập vào không phải là định dạng số!")
            
    return render(request, 'admin/admin_kho_form.html', {'kho': kho})


@login_required
def admin_xoa_kho(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    kho = get_object_or_404(NhaCungCap, id=id)
    
    try:
        ten_kho = kho.ten_ncc
        kho.delete()
        messages.success(request, f"Đã xóa thành công kho: {ten_kho}!")
    except ProtectedError:
        # Lỗi này văng ra khi bạn xóa 1 kho mà kho đó đang có dữ liệu lô hàng, nhập kho liên kết với nó
        messages.error(request, "Không thể xóa! Kho này đang chứa dữ liệu lịch sử nhập hàng. Vui lòng chỉ chỉnh sửa trạng thái thay vì xóa.")
    except Exception as e:
        messages.error(request, f"Lỗi không xác định khi xóa: {e}")
        
    return redirect('admin_khos')

# ========================================================
# 7. QUẢN LÝ DANH MỤC TRƯNG BÀY (CMS CATALOG)
# ========================================================
@login_required
def admin_san_pham(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Bạn không có quyền truy cập trang này!")
        return redirect('trang_chu')

    ds_sp = SanPham.objects.select_related('danh_muc').all().order_by('-id')
    ds_danh_muc = DanhMuc.objects.all()

    context = {
        'ds_sp': ds_sp,
        'ds_danh_muc': ds_danh_muc,
    }
    return render(request, 'admin/admin_san_pham.html', context)

@login_required
def admin_luu_san_pham(request):
    if request.method == 'POST':
        sp_id = request.POST.get('sp_id')
        ten_sp = request.POST.get('ten_san_pham') or request.POST.get('ten_sp') # Hỗ trợ tên field từ form HTML cũ/mới
        danh_muc_id = request.POST.get('danh_muc_id')
        gia_tham_khao = float(request.POST.get('gia_ban', request.POST.get('gia_tham_khao', 0))) 
        mo_ta = request.POST.get('mo_ta', '')
        anh_sp = request.FILES.get('anh_san_pham') or request.FILES.get('anh_sp') 

        try:
            danh_muc = DanhMuc.objects.get(id=danh_muc_id)
            
            if sp_id:  
                sp = SanPham.objects.get(id=sp_id)
                sp.ten_sp = ten_sp
                sp.danh_muc = danh_muc
                sp.gia_tham_khao = gia_tham_khao
                sp.mo_ta = mo_ta
                if anh_sp: 
                    sp.anh_sp = anh_sp
                sp.save()
                messages.success(request, f"Đã cập nhật trưng bày: {ten_sp}!")
            else:  
                SanPham.objects.create(
                    ten_sp=ten_sp,
                    danh_muc=danh_muc,
                    gia_tham_khao=gia_tham_khao,
                    mo_ta=mo_ta,
                    anh_sp=anh_sp
                )
                messages.success(request, f"Đã thêm vào trưng bày: {ten_sp}!")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")

    return redirect('admin_san_pham')

@login_required
def admin_xoa_san_pham(request, sp_id):
    try:
        SanPham.objects.filter(id=sp_id).delete()
        messages.success(request, "Đã xóa khỏi danh mục trưng bày!")
    except Exception as e:
        messages.error(request, f"Lỗi: {e}")
    return redirect('admin_san_pham')


@login_required
def admin_luu_danh_muc(request):
    if request.method == 'POST':
        ten_dm = request.POST.get('ten_danh_muc')
        if ten_dm:
            DanhMuc.objects.create(ten_dm=ten_dm)
            messages.success(request, f"Đã thêm nhóm trưng bày: {ten_dm}")
    return redirect('admin_san_pham')

@login_required
def admin_xoa_danh_muc(request, dm_id):
    try:
        DanhMuc.objects.filter(id=dm_id).delete()
        messages.success(request, "Đã xóa nhóm trưng bày!")
    except Exception as e:
        messages.error(request, f"Lỗi: {e}")
    return redirect('admin_san_pham')

# ========================================================
# 8. QUẢN LÝ BỒN CHỨA (TANK MANAGEMENT)
# ========================================================
@login_required
def admin_bons(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Bạn không có quyền truy cập trang này!")
        return redirect('trang_chu')

    ds_bon = BonChua.objects.select_related('tram').all().order_by('tram__ten_tram', 'ten_bon')
    ds_tram = TramXang.objects.all()

    context = {'ds_bon': ds_bon, 'ds_tram': ds_tram}
    return render(request, 'admin/admin_bons.html', context)

@login_required
def admin_luu_bon(request):
    if request.method == 'POST':
        bon_id = request.POST.get('bon_id')
        tram_id = request.POST.get('tram_id')
        ten_bon = request.POST.get('ten_bon')
        loai_nhien_lieu = request.POST.get('loai_nhien_lieu')
        suc_chua_toi_da = float(request.POST.get('suc_chua_toi_da', 0))

        try:
            tram = TramXang.objects.get(id=tram_id)
            if bon_id: 
                bon = BonChua.objects.get(id=bon_id)
                bon.tram = tram
                bon.ten_bon = ten_bon
                bon.loai_nhien_lieu = loai_nhien_lieu
                bon.suc_chua_toi_da = suc_chua_toi_da
                bon.save()
                messages.success(request, f"Đã cập nhật {ten_bon}!")
            else: 
                BonChua.objects.create(
                    tram=tram, ten_bon=ten_bon, loai_nhien_lieu=loai_nhien_lieu,
                    suc_chua_toi_da=suc_chua_toi_da, muc_hien_tai=0 
                )
                messages.success(request, f"Đã xây thêm {ten_bon}!")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")

    return redirect('admin_bons')

@login_required
def admin_xoa_bon(request, bon_id):
    try:
        bon = BonChua.objects.get(id=bon_id)
        bon.delete()
        messages.success(request, f"Đã đập bỏ bồn chứa!")
    except Exception as e:
        messages.error(request, f"Không thể xóa: {e}")
    return redirect('admin_bons')

# ==========================================
# 9. CRUD: NHÂN SỰ
# ==========================================
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password

# Lấy model User hiện tại của hệ thống (phòng trường hợp bạn custom User)
User = get_user_model()

@login_required
def admin_danh_sach_nhan_su(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
    
    # BỎ EXCLUDE: Lấy tất cả nhân sự để Admin có thể tự xem thông tin của mình
    ds_nhan_su = User.objects.all().order_by('-date_joined')
    return render(request, 'admin/admin_nhan_sus.html', {'ds_nhan_su': ds_nhan_su})


@login_required
def admin_them_nhan_su(request):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    ds_tram = TramXang.objects.filter(trang_thai='hoat_dong')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        full_name = request.POST.get('full_name')
        phone = request.POST.get('phone')
        role = request.POST.get('role')
        tram_id = request.POST.get('tram_id')

        # Logic ép kiểu: Nếu là admin thì không thuộc trạm nào (Văn phòng)
        if role == 'admin':
            tram_id = None

        try:
            # Kiểm tra xem username đã tồn tại chưa
            if User.objects.filter(username=username).exists():
                messages.error(request, f"Tài khoản '{username}' đã tồn tại. Vui lòng chọn tên khác!")
                return render(request, 'admin/admin_nhan_su_form.html', {'ds_tram': ds_tram})

            # Khởi tạo User mới
            user = User(
                username=username,
                full_name=full_name,
                phone=phone,
                role=role,
                tram_xang_id=tram_id
            )
            # Bắt buộc phải dùng set_password để mã hóa mật khẩu
            user.set_password(password)
            user.save()
            
            messages.success(request, f"Đã tạo thành công tài khoản: {username}")
            return redirect('admin_nhan_sus')
            
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
            
    return render(request, 'admin/admin_nhan_su_form.html', {'ds_tram': ds_tram})


@login_required
def admin_sua_nhan_su(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    nv = get_object_or_404(User, id=id)
    ds_tram = TramXang.objects.filter(trang_thai='hoat_dong')
    
    if request.method == 'POST':
        try:
            nv.full_name = request.POST.get('full_name')
            nv.phone = request.POST.get('phone')
            
            # Chỉ cho phép đổi Role nếu không phải đang sửa chính mình
            # (Tránh trường hợp Admin tự giáng cấp bản thân thành nhân viên)
            if request.user.id != nv.id:
                nv.role = request.POST.get('role')
                # Check trạng thái hoạt động (bật/tắt)
                nv.is_active = request.POST.get('is_active') == 'on'

            # Xử lý Nơi làm việc
            tram_id = request.POST.get('tram_id')
            if nv.role == 'admin':
                nv.tram_xang_id = None
            else:
                nv.tram_xang_id = tram_id

            # Xử lý đổi mật khẩu (nếu có nhập)
            new_password = request.POST.get('password')
            if new_password: 
                nv.set_password(new_password)
                
            nv.save()
            messages.success(request, f"Đã cập nhật hồ sơ của {nv.full_name}!")
            return redirect('admin_nhan_sus')
            
        except Exception as e:
            messages.error(request, f"Lỗi cập nhật: {e}")
            
    return render(request, 'admin/admin_nhan_su_form.html', {'nv': nv, 'ds_tram': ds_tram})


@login_required
def admin_xoa_nhan_su(request, id):
    if request.user.role != 'admin' and not request.user.is_superuser: 
        return redirect('trang_chu')
        
    nv = get_object_or_404(User, id=id)
    
    # 1. Chốt chặn bảo mật: KHÔNG CHO XÓA ADMIN / SUPERUSER
    if nv.role == 'admin' or nv.is_superuser:
        messages.error(request, "CẢNH BÁO BẢO MẬT: Bạn không thể xóa tài khoản Quản trị viên hệ thống!")
        return redirect('admin_nhan_sus')
        
    # 2. Chốt chặn bảo mật: Không cho tự xóa chính mình
    if request.user.id == nv.id:
        messages.warning(request, "Bạn không thể tự xóa tài khoản đang đăng nhập của chính mình!")
        return redirect('admin_nhan_sus')

    try:
        # 3. Kiểm tra ràng buộc nghiệp vụ (Lịch sử bán hàng)
        # Sửa chữ 'HoaDon' thành bảng giao dịch/xuất kho tương ứng của bạn nếu cần
        if hasattr(nv, 'hoadon_set') and nv.hoadon_set.exists(): 
            messages.error(request, f"Không thể xóa {nv.full_name} vì nhân viên này đã có lịch sử xuất hóa đơn/nhập hàng!")
            messages.info(request, "Gợi ý: Hãy bấm 'Sửa' và 'Khóa tài khoản' thay vì Xóa.")
        else: 
            t = nv.full_name
            nv.delete()
            messages.success(request, f"Đã xóa vĩnh viễn tài khoản nhân sự: {t}")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa nhân sự: {e}")
        
    return redirect('admin_nhan_sus')

# ==========================================
# 10. CRUD: TIN TỨC & BANNER
# ==========================================
@login_required
def admin_tin_tuc(request):
    if request.user.role != 'admin' and not request.user.is_superuser: return redirect('trang_chu')
    ds_tin = TinTuc.objects.all().order_by('-ngay_dang')
    return render(request, 'admin/admin_tin_tuc.html', {'ds_tin': ds_tin})

@login_required
def admin_tin_tuc_form(request, tin_id=None):
    if request.user.role != 'admin' and not request.user.is_superuser: return redirect('trang_chu')
    tin_hien_tai = get_object_or_404(TinTuc, id=tin_id) if tin_id else None

    if request.method == 'POST':
        try:
            if not tin_hien_tai: tin_hien_tai = TinTuc() 
            tin_hien_tai.tieu_de = request.POST.get('tieu_de')
            tin_hien_tai.tom_tat = request.POST.get('tom_tat')
            tin_hien_tai.noi_dung = request.POST.get('noi_dung')
            
            anh_bia = request.FILES.get('anh_bia') 
            if anh_bia: tin_hien_tai.anh_bia = anh_bia
            tin_hien_tai.save()
            messages.success(request, "Lưu bài viết thành công!")
            return redirect('admin_tin_tuc')
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")

    return render(request, 'admin/admin_tin_tuc_form.html', {'tin': tin_hien_tai})

@login_required
def admin_xoa_tin_tuc(request, tin_id):
    if request.user.role == 'admin' or request.user.is_superuser:
        get_object_or_404(TinTuc, id=tin_id).delete()
        messages.success(request, "Đã xóa bài viết!")
    return redirect('admin_tin_tuc')

# ==========================================
# 11. HỘP THƯ (CMS INBOX)
# ==========================================
def gui_yeu_cau_b2b(request):
    if request.method == 'POST':
        try:
            DoiTacB2B.objects.create(
                ten_cong_ty=request.POST.get('ten_cong_ty'), so_dien_thoai=request.POST.get('so_dien_thoai'),
                email=request.POST.get('email'), nhu_cau=request.POST.get('nhu_cau')
            )
            messages.success(request, "Đã gửi yêu cầu hợp tác! GSMS sẽ liên hệ với bạn sớm nhất.")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
    return redirect('doi_tac')

def nop_ho_so(request):
    if request.method == 'POST':
        try:
            HoSoUngVien.objects.create(
                vi_tri_ung_tuyen=request.POST.get('vi_tri'), ho_ten=request.POST.get('ho_ten'),
                so_dien_thoai=request.POST.get('so_dien_thoai'), email=request.POST.get('email'),
                file_cv=request.FILES.get('file_cv') 
            )
            messages.success(request, "Nộp hồ sơ thành công! Chúc bạn may mắn.")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
    return redirect('tuyen_dung')

@login_required
def admin_inbox(request):
    if request.user.role not in ['admin', 'ke_toan'] and not request.user.is_superuser:
        return redirect('trang_chu')

    if request.method == 'POST':
        loai = request.POST.get('loai')
        item_id = request.POST.get('id')
        hanh_dong = request.POST.get('hanh_dong')
        
        try:
            if loai == 'b2b':
                obj = DoiTacB2B.objects.get(id=item_id)
                obj.trang_thai = hanh_dong
                obj.save()
                messages.success(request, "Đã cập nhật trạng thái hợp tác B2B!")
                
            elif loai == 'ung_vien':
                obj = HoSoUngVien.objects.get(id=item_id)
                obj.trang_thai = hanh_dong
                obj.save()
                messages.success(request, "Đã cập nhật trạng thái hồ sơ ứng viên!")
                
            elif loai == 'gop_y':
                obj = LienHeGopY.objects.get(id=item_id)
                obj.da_xu_ly = (hanh_dong == 'xong')
                obj.save()
                if hanh_dong == 'xong':
                    messages.success(request, "Đã đánh dấu xử lý xong góp ý!")
                else:
                    messages.warning(request, "Đã đưa góp ý về lại trạng thái chờ.")
                    
            elif loai == 'phan_hoi_gmail':
                gop_y = LienHeGopY.objects.get(id=item_id)
                noi_dung_tra_loi = request.POST.get('noi_dung_tra_loi')
                
                # Kiểm tra xem khách có email không
                email_khach = getattr(gop_y, 'email', None) 
                
                if email_khach:
                    subject = f"GSMS - Phản hồi về nội dung: {gop_y.tieu_de}"
                    message = f"Chào {gop_y.ho_ten},\n\n{noi_dung_tra_loi}\n\nTrân trọng,\nBan điều hành GSMS."
                    
                    try:
                        # 1. Thực thi gửi mail thật
                        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email_khach], fail_silently=False)
                        
                        # 2. ĐÂY LÀ KHÚC QUAN TRỌNG NHẤT: Đổi trạng thái và Lưu lại
                        gop_y.da_xu_ly = True 
                        gop_y.save()
                        
                        messages.success(request, f"Đã gửi mail phản hồi và tự động đánh dấu 'Đã ghi nhận'!")
                    except Exception as e:
                        messages.error(request, f"Hệ thống gửi mail thất bại. Lỗi: {e}")
                else:
                    messages.error(request, "Không thể gửi! Khách hàng này không để lại địa chỉ Email.")

            elif loai == 'danh_gia':
                obj = DanhGiaTram.objects.get(id=item_id)
                if hanh_dong == 'duyet':
                    obj.da_duyet = True
                    obj.save()
                    messages.success(request, "Đã duyệt công khai bài đánh giá!")
                elif hanh_dong == 'an_bai':
                    obj.da_duyet = False
                    obj.save()
                    messages.warning(request, "Đã ẩn bài đánh giá khỏi trang web!")
                elif hanh_dong == 'xoa':
                    obj.delete() 
                    messages.success(request, "Đã xóa vĩnh viễn đánh giá rác!")
                    
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
            
        return redirect('admin_inbox')

    context = {
        'ds_b2b': DoiTacB2B.objects.all().order_by('-ngay_gui'),
        'ds_ung_vien': HoSoUngVien.objects.all().order_by('-ngay_nop'),
        'ds_gop_y': LienHeGopY.objects.all().order_by('da_xu_ly', '-ngay_gui'),
        'ds_danh_gia': DanhGiaTram.objects.all().order_by('da_duyet', '-ngay_gui'), 
    }
    return render(request, 'admin/admin_inbox.html', context)

def gui_danh_gia(request, sp_id):
    if request.method == 'POST':
        try:
            sp = get_object_or_404(SanPham, id=sp_id)
            DanhGiaTram.objects.create(
                san_pham=sp,
                ten_khach_hang=request.POST.get('ten_khach_hang'),
                so_sao=request.POST.get('so_sao', 5), 
                noi_dung=request.POST.get('noi_dung')
            )
            messages.success(request, "Cảm ơn bạn! Đánh giá đã được gửi và đang chờ Ban quản trị kiểm duyệt.")
        except Exception as e:
            messages.error(request, f"Lỗi gửi đánh giá: {e}")
            
    return redirect(request.META.get('HTTP_REFERER', 'san_pham'))
def gui_danh_gia_tram(request):
    if request.method == 'POST':
        tram_id = request.POST.get('tram_id')
        ten_khach_hang = request.POST.get('ten_khach_hang')
        so_sao = request.POST.get('so_sao')
        noi_dung = request.POST.get('noi_dung')

        try:
            tram = TramXang.objects.get(id=tram_id)
            DanhGiaTram.objects.create(
                tram=tram,
                ten_khach_hang=ten_khach_hang,
                so_sao=so_sao,
                noi_dung=noi_dung
                # da_duyet mặc định là False, Admin sẽ duyệt trong hộp thư
            )
            messages.success(request, "Cảm ơn bạn đã đánh giá! Hệ thống đã ghi nhận.")
        except Exception as e:
            messages.error(request, f"Lỗi gửi đánh giá: {e}")
            
    return redirect('trang_chu')